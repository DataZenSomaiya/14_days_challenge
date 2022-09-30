from openpyxl import load_workbook
  
# loading the excel sheet
book = load_workbook('attendance.xlsx')
  
# Choose the sheet
sheet = book['Sheet1']
  
# counting number of rows / students
r = sheet.max_row
  
# variable for looping for input
resp = 1
  
# counting number of columns / subjects
c = sheet.max_column
    
def savefile():
    book.save('attendance.xlsx')
    print("saved!")
  
while resp == 1:
    print('### Welcome to Attendance Tracker ###')
    print("1--->EC\n2--->PP\n3--->AM")
  
    # enter the corresponding number
    y = int(input("Enter the subject :"))
    if (y!=3 and y!=2 and y!=1):
        print('Invalid choice')
        continue
    # no.of.absentees for that subject
    no_of_absentees = int(input('Enter the number of absentees :'))
  
    if(no_of_absentees > 1):
        x = list(map(int, (input('Enter the roll numbers of absentees :').split( ))))
    else:
        x = [int(input('Enter the roll number of the absentee :'))]
  
    for student in x:
  
        for i in range(2, r+1):
  
            if y==1:
                if sheet.cell(row=i, column=1).value is student:
                    m = sheet.cell(row=i, column=3).value
                    m = m+1
                    sheet.cell(row=i, column=3).value = m
                    savefile()
  
            elif y == 2:
                if sheet.cell(row=i, column=1).value is student:
                    m = sheet.cell(row=i, column=4).value
                    m = m+1
                    sheet.cell(row=i, column=4).value = m
                    savefile()
  
            elif y == 3:
                if sheet.cell(row=i, column=1).value is student:
                    m = sheet.cell(row=i, column=5).value
                    m = m+1
                    sheet.cell(row=i, column=5).value = m
                    savefile()

    resp = int(input('Select another subject ?  1---->YES  0--->NO :'))
    if resp==0:
        print("END!")